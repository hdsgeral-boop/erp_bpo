import sys
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def format_kz(val):
    if val is None:
        val = 0.0
    return f"{val:,.2f} Kz".replace(",", " ").replace(".", ",")

def generate_inss(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapa INSS"
    ws.views.sheetView[0].showGridLines = True

    # Styling
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    border_total = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    headers = [
        "Colaborador", "Nº INSS", "Sal. Base", "Subs. Adic.", 
        "Total Base", "Trab. (3%)", "Empresa (8%)", "Total (11%)"
    ]

    # Write Header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_num == 2 else ("right" if col_num >= 3 else "left"), vertical="center")

    tot_base_sal = 0.0
    tot_subs_adic = 0.0
    tot_total_base = 0.0
    tot_trab_3 = 0.0
    tot_emp_8 = 0.0
    tot_total_11 = 0.0

    row_idx = 2
    for item in data.get("items", []):
        base_sal = float(item.get("base_salary", 0.0))
        inss_base = float(item.get("inss_base", base_sal))
        subs_adic = max(0.0, inss_base - base_sal)
        trab_3 = float(item.get("inss_employee", inss_base * 0.03))
        emp_8 = float(item.get("inss_company", inss_base * 0.08))
        total_11 = trab_3 + emp_8

        tot_base_sal += base_sal
        tot_subs_adic += subs_adic
        tot_total_base += inss_base
        tot_trab_3 += trab_3
        tot_emp_8 += emp_8
        tot_total_11 += total_11

        row_vals = [
            item.get("name", ""),
            item.get("inss", "N/A"),
            format_kz(base_sal),
            format_kz(subs_adic),
            format_kz(inss_base),
            format_kz(trab_3),
            format_kz(emp_8),
            format_kz(total_11)
        ]

        for col_num, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=val)
            cell.font = font_regular
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="center" if col_num == 2 else ("right" if col_num >= 3 else "left"), vertical="center")

        row_idx += 1

    # Total Row
    total_row = [
        "TOTAL:", "", format_kz(tot_base_sal), format_kz(tot_subs_adic),
        format_kz(tot_total_base), format_kz(tot_trab_3), format_kz(tot_emp_8), format_kz(tot_total_11)
    ]
    for col_num, val in enumerate(total_row, 1):
        cell = ws.cell(row=row_idx, column=col_num, value=val)
        cell.font = font_bold
        cell.border = border_total
        cell.alignment = Alignment(horizontal="center" if col_num == 2 else ("right" if col_num >= 3 else "left"), vertical="center")

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_path)

def generate_irt(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapa IRT AGT"
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)

    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    border_total = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    company_name = data.get("company_name", "ERP CONSULVOLT - SISTEMA INTEGRADO DE GESTÃO")
    ws.cell(row=1, column=1, value=f"{company_name} - Modelo Oficial IRT AGT").font = font_title

    headers = [
        "Ord.", "NIF", "INSS", "Nome Completo", "Província", "Município", "Período",
        "S. Alim", "S. Transp", "S. férias", "Sal. Base", "Total Bruto",
        "S. Social (3%)", "Outras Ded.", "Matéria Col.", "Parcela Fixa",
        "Taxa %", "Excesso", "Imp. Devido", "Imp. Retido", "A Pagar/Reemb."
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_num in [1,2,3,5,6,7,17] else ("right" if col_num >= 8 else "left"), vertical="center")

    tot_alim = 0.0
    tot_transp = 0.0
    tot_ferias = 0.0
    tot_base = 0.0
    tot_bruto = 0.0
    tot_inss = 0.0
    tot_materia = 0.0
    tot_fixa = 0.0
    tot_devido = 0.0
    tot_retido = 0.0

    periodo = data.get("reference", "07/2026")
    row_idx = 3

    for idx, item in enumerate(data.get("items", []), 1):
        base_sal = float(item.get("base_salary", 0.0))
        additions = float(item.get("additions", 0.0))
        s_alim = float(item.get("subsidy_meal", 30000.0 if additions >= 30000 else additions))
        s_transp = float(item.get("subsidy_transport", max(0.0, additions - s_alim)))
        s_ferias = max(0.0, additions - s_alim - s_transp)

        total_bruto = base_sal + additions
        inss_3 = float(item.get("inss_employee", 0.0))
        
        # Isenções
        taxable_meal = max(0.0, s_alim - 30000.0)
        taxable_transp = max(0.0, s_transp - 30000.0)
        materia_col = max(0.0, (base_sal + taxable_meal + taxable_transp + s_ferias) - inss_3)
        
        irt_retido = float(item.get("irt", 0.0))
        
        tot_alim += s_alim
        tot_transp += s_transp
        tot_ferias += s_ferias
        tot_base += base_sal
        tot_bruto += total_bruto
        tot_inss += inss_3
        tot_materia += materia_col
        tot_retido += irt_retido
        tot_devido += irt_retido

        row_vals = [
            idx,
            item.get("nif", "N/A"),
            item.get("inss", "N/A"),
            item.get("name", ""),
            item.get("province", "Luanda"),
            item.get("municipality", "---"),
            periodo,
            format_kz(s_alim),
            format_kz(s_transp),
            format_kz(s_ferias),
            format_kz(base_sal),
            format_kz(total_bruto),
            format_kz(inss_3),
            "0,00 Kz",
            format_kz(materia_col),
            "0,00 Kz",
            "0%",
            format_kz(materia_col),
            format_kz(irt_retido),
            format_kz(irt_retido),
            "0"
        ]

        for col_num, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=val)
            cell.font = font_regular
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="center" if col_num in [1,2,3,5,6,7,17] else ("right" if col_num >= 8 else "left"), vertical="center")

        row_idx += 1

    # Total Row
    total_row = [
        "TOTAL:", "", "", "", "", "", "",
        format_kz(tot_alim), format_kz(tot_transp), format_kz(tot_ferias),
        format_kz(tot_base), format_kz(tot_bruto), format_kz(tot_inss),
        "0,00 Kz", format_kz(tot_materia), "0,00 Kz", "", "",
        format_kz(tot_devido), format_kz(tot_retido), "0"
    ]
    for col_num, val in enumerate(total_row, 1):
        cell = ws.cell(row=row_idx, column=col_num, value=val)
        cell.font = font_bold
        cell.border = border_total
        cell.alignment = Alignment(horizontal="center" if col_num in [1,2,3,5,6,7,17] else ("right" if col_num >= 8 else "left"), vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    wb.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        sys.exit(1)
    
    fmt = sys.argv[1]
    json_path = sys.argv[2]
    out_path = sys.argv[3]

    with open(json_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)

    if fmt == "inss":
        generate_inss(payload, out_path)
    elif fmt == "irt":
        generate_irt(payload, out_path)
